# -*- coding: utf-8 -*-
"""
Script to convert networks from PyPSA-Eur-Sec v0.0.2 to data format used in the
IPCC AR6 database
"""

import pypsa
import openpyxl
import pandas as pd

#original IPCC file, official template
template_path = "global_sectoral/IPCC_AR6_WG3_Global_sectoral_Pathways_scenario_template_v3.1_20221027.xlsx"
#use the official template

# Metadata
model = "PyPSA-Eur-Sec 0.0.2"
model_version = "0.0.6"
literature_reference = "Pedersen, T. T., Gøtske, E. K., Dvorak, A., Andresen, G. B., & Victoria, M. (2022). Long-term implications of reduced gas imports on the decarbonization of the European energy system. Joule, 6(7), 1566-1580."
climate_target = 21 # CO2 budget

wind_split = ['DE', 'ES', 'FI', 'FR', 'GB', 'IT', 'NO', 'PL', 'RO', 'SE']

scenarios={'Base_1.5':'postnetworks/elec_s370_37m_lv1.0__3H-T-H-B-I-solar+p3-dist1-cb25.7ex0_',
            'Gaslimit_1.5':'postnetworks/elec_s370_37m_lv1.0__3H-T-H-B-I-solar+p3-dist1-cb25.7ex0-gasconstrained_',
            'Base_2':'postnetworks/elec_s370_37m_lv1.0__3H-T-H-B-I-solar+p3-dist1-cb73.9ex0_',
            'Gaslimit_2': 'postnetworks/elec_s370_37m_lv1.0__3H-T-H-B-I-solar+p3-dist1-cb73.9ex0-gasconstrained_',
                     }


output_folder = 'results/'

years = [2020, 2025, 2030, 2035, 2040, 2045, 2050]

countries = ['AT','BE','BG','CH','CZ','DE','DK','EE','ES','FI','FR','GB','GR','HR',
             'HU','IT','LT','LU','LV','NO','PL','PT','RO','SE','SI','SK','IE', 'NL',
             #'RS','BA'
             ] 

iso2name={'AT':'Austria',
          'BE':'Belgium',
          'BG':'Bulgaria',
          'CH':'Switzerland',
          'CZ':'Czech Republic',
          'DE':'Germany',
          'DK':'Denmark',
          'EE':'Estonia',
          'ES':'Spain',
          'FI':'Finland',
          'FR':'France',
          'GB':'United Kingdom',
          'GR':'Greece',
          'HR':'Croatia',
          'HU':'Hungary',
          'IT':'Italy',
          'LT':'Lithuania',
          'LU':'Luxembourg',
          'LV':'Latvia',
          'NO':'Norway',
          'PL':'Poland',
          'PT':'Portugal',
          'RO':'Romania',
          'SE':'Sweden',
          'SI':'Slovenia',
          'SK':'Slovakia',
          'IE':'Ireland',
          'NL':'The Netherlands',}

for scenario in scenarios:
    #one excel file per scenario
    file = openpyxl.load_workbook(template_path)
    ds = file['data'] #data sheet
    for year in years:
        n = pypsa.Network(f"{scenarios[scenario]}{year}.nc")
        costs = pd.read_csv(f"costs/costs_{year}.csv", index_col=[0,1])
        
        col=[c for c in ds[1] if c.value==year][0].column
        
        for i,country in enumerate(countries):
            if year == 2020:
                #one datasheet per country including information from different years
                target = file.copy_worksheet(file['data'])
                target.title ='data' + str(i)
            ds = file['data' + str(i)] 
            var={}
            
            """
            Capacity : Solar PV, onshore and offshore wind
            """
            #MW -> GW
            var['Capacity|Electricity|Solar|PV'] =0.001*(n.generators.p_nom_opt.filter(like ='solar').filter(like =country).sum()-
                                                          n.generators.p_nom_opt.filter(like ='solar thermal').filter(like =country).sum())
            var['Capacity|Electricity|Solar'] = var['Capacity|Electricity|Solar|PV']
            var['Capacity |Electricity|Solar|PV|Rooftop PV'] = 0.001*n.generators.p_nom_opt.filter(like ='solar rooftop').filter(like =country).sum()
            var['Capacity|Electricity|Solar|PV|Utility-scale PV'] = var['Capacity|Electricity|Solar|PV'] - var['Capacity |Electricity|Solar|PV|Rooftop PV']

            var['Capacity|Electricity|Wind|Onshore']=0.001*n.generators.p_nom_opt.filter(like ='onwind').filter(like =country).sum() 
            var['Capacity|Electricity|Wind|Offshore']=0.001*n.generators.p_nom_opt.filter(like ='offwind').filter(like =country).sum() 
            var['Capacity|Electricity|Wind']=var['Capacity|Electricity|Wind|Onshore']+var['Capacity|Electricity|Wind|Offshore']


            """
            Capacity : Nuclear, Coal, Lignite, OCGT, CCGT, Biomass
            """
            #MW -> GW
            var['Capacity|Electricity|Nuclear'] =0.001*((n.links.efficiency.filter(like ='nuclear').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='nuclear').filter(like =country)).sum())
            
            var['Capacity|Electricity|Coal|w/o CCS'] = ((n.links.efficiency.filter(like ='coal').filter(like =country)
                *n.links.p_nom_opt.filter(like ='coal').filter(like =country)).sum())
            var['Capacity|Electricity|Coal|w/o CCS'] += 0.001*((n.links.efficiency.filter(like ='lignite').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='lignite').filter(like =country)).sum())
            var['Capacity|Electricity|Coal'] =var['Capacity|Electricity|Coal|w/o CCS'] 
            
            var['Capacity|Electricity|Gas'] = 0.001*((n.links.efficiency.filter(like ='OCGT').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='OCGT').filter(like =country)).sum())
            var['Capacity|Electricity|Gas'] += 0.001*((n.links.efficiency.filter(like ='CCGT').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='CCGT').filter(like =country)).sum())
            var['Capacity|Electricity|Gas'] += 0.001*((n.links.efficiency.filter(like ='gas CHP').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='gas CHP').filter(like =country)).sum())
            var['Capacity|Electricity|Gas|w/ CCS'] = 0.001*((n.links.efficiency.filter(like ='gas CHP CC').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='gas CHP CC').filter(like =country)).sum() )                                                 
            var['Capacity|Electricity|Gas|w/o CCS'] = (var['Capacity|Electricity|Gas'] -
                                         var['Capacity|Electricity|Gas|w/ CCS']  )                                              
                                                              
            var['Capacity|Electricity|Biomass'] = 0.001*((n.links.efficiency.filter(like ='solid biomass CHP').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='solid biomass CHP').filter(like =country)).sum())
            var['Capacity|Electricity|Biomass|w/ CCS']= 0.001*((n.links.efficiency.filter(like ='solid biomass CHP CC').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='solid biomass CHP CC').filter(like =country)).sum())   
            var['Capacity|Electricity|Biomass|w/o CCS'] = (var['Capacity|Electricity|Biomass']    
              -var['Capacity|Electricity|Biomass|w/ CCS'])
            
            """
            Capacity : hydro (reservoir, ror)
            """
            #MW -> GW
            var['Capacity|Electricity|Hydro'] = 0.001*n.generators.p_nom_opt.filter(like ='ror').filter(like =country).sum()
            var['Capacity|Electricity|Hydro'] += 0.001*n.storage_units.p_nom_opt.filter(like ='hydro').filter(like =country).sum()
            """
            Electricity : Solar PV, onshore and offshore wind
            """
            #MWh -> EJ
            var['Secondary Energy|Electricity|Solar|PV'] =3.6e-9*(n.generators_t.p.filter(like ='solar').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Solar'] = var['Secondary Energy|Electricity|Solar|PV']
            var['Secondary Energy|Electricity|Solar|PV|Rooftop PV'] = 3.6e-9*(n.generators_t.p.filter(like ='solar rooftop').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Solar|PV|Utility-scale PV'] = var['Secondary Energy|Electricity|Solar|PV'] - var['Secondary Energy|Electricity|Solar|PV|Rooftop PV']

            var['Secondary Energy|Electricity|Wind|Onshore'] = 3.6e-9*(n.generators_t.p.filter(like ='onwind').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Wind|Offshore'] = 3.6e-9*(n.generators_t.p.filter(like ='offwind').filter(like =country).sum().sum())
            
            var['Secondary Energy|Electricity|Wind'] = var['Secondary Energy|Electricity|Wind|Onshore'] + var['Secondary Energy|Electricity|Wind|Offshore']
            
            """
            Electricity : Nuclear, Coal, Lignite, OCGT, CCGT, biomass
            """
            #MWh -> EJ
            var['Secondary Energy|Electricity|Nuclear'] = -3.6e-9*(n.links_t.p1.filter(like ='nuclear').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Coal|w/o CCS'] =- 3.6e-9*(n.links_t.p1.filter(like ='coal').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Coal|w/o CCS'] += -3.6e-9*(n.links_t.p1.filter(like ='lignite').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Gas|w/o CCS'] = -3.6e-9*(n.links_t.p1.filter(like ='OCGT').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Gas|w/o CCS'] += -3.6e-9*(n.links_t.p1.filter(like ='CCGT').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Gas|w/o CCS'] += -3.6e-9*(n.links_t.p1.filter(like ='gas CHP').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Biomass|w/o CCS'] = -3.6e-9*(n.links_t.p1.filter(like ='biomass CHP').filter(like =country).sum().sum())
                                                                 
            var['Secondary Energy|Electricity|Gas|w/ CCS'] = -3.6e-9*(n.links_t.p1.filter(like ='gas CHP CC').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Biomass|w/ CCS'] = -3.6e-9*(n.links_t.p1.filter(like ='biomass CHP CC').filter(like =country).sum().sum())           
            var['Secondary Energy|Electricity|Gas|w/o CCS'] -= var['Secondary Energy|Electricity|Gas|w/ CCS']
            var['Secondary Energy|Electricity|Biomass|w/o CCS'] -= var['Secondary Energy|Electricity|Biomass|w/ CCS']
            """
            Electricity : Hydro (reservoir, ror)
            """
            #MWh -> EJ
            var['Secondary Energy|Electricity|Hydro'] = 3.6e-9*(n.storage_units_t.p.filter(like ='hydro').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Hydro'] += 3.6e-9*(n.generators_t.p.filter(like ='ror').filter(like =country).sum().sum())

            """
            Capacity : storage (PHS, battery, H2 storage)
            """
            #MWh to GWh
            var['Capacity|Electricity|Storage|Pumped Hydro Storage'] = 0.001*(n.storage_units.p_nom_opt.filter(like ='PHS').filter(like =country).sum())
            var['Capacity|Electricity|Storage|Battery Capacity|Utility-scale Battery '] = 0.001*((n.links.efficiency.filter(like ='battery charger')
                 *n.links.p_nom_opt.filter(like ='battery charger')).sum()-(n.links.efficiency.filter(like ='home battery charger')
                 *n.links.p_nom_opt.filter(like ='home battery charger')).sum())
            var['Capacity|Electricity|Storage|Battery Capacity'] = 0.001*((n.links.efficiency.filter(like ='battery charger')
                 *n.links.p_nom_opt.filter(like ='battery charger')).sum())
            #var['Capacity|Electricity|Storage|Hydrogen Storage Capacity|overground'] = 0.001 *(n.stores.e_nom_opt.filter(like ='H2').filter(like =country).sum()/168) #assume one week charge time for H2 storage
            #var['Capacity|Electricity|Storage|Hydrogen Storage Capacity|underground'] = 0.001 *n.stores.e_nom_opt[country + ' H2 Store underground'] if country + ' H2 Store underground' in n.stores.index else 0
            var['Capacity|Electricity|Storage|Hydrogen Storage Capacity'] = 0.001 *(n.stores.e_nom_opt.filter(like ='H2').filter(like =country).sum()/168)
            var['Capacity|Electricity|Storage Capacity'] = ( var['Capacity|Electricity|Storage|Pumped Hydro Storage']
                                                            +  var['Capacity|Electricity|Storage|Battery Capacity']
                                                            + var['Capacity|Electricity|Storage|Hydrogen Storage Capacity'])
            
            """
            Capacity : heat pumps, heat resistors, Sabatier (synthetic gas)
            """
            # ELectric capacity
            # MW to Gw
            var['Capacity|Heating|Heat pumps'] = 0.001*((n.links_t.efficiency.filter(like ='heat pump').filter(like =country).mean()
                 *n.links.p_nom_opt.filter(like ='heat pump').filter(like =country)).sum())
            var['Capacity|Heating|Electric boilers'] = 0.001*((n.links.efficiency.filter(like ='resistive heater').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='resistive heater').filter(like =country)).sum())
            var['Capacity|Gas|Synthetic'] = 0.001*((n.links.efficiency.filter(like ='Sabatier').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='Sabatier').filter(like =country)).sum())
            
            """
            Final Energy (heating) : heat pumps, heat resistors, Sabatier (synthetic gas)
            """
            #MWh -> EJ
            #50/50 services/domestic
            var['Final Energy|Residential and Commercial|Commercial|Heating|Heat pumps'] = - 3.6e-9 * 0.5 * (n.links_t.p1.filter(like ='heat pump').filter(like='residentioal').filter(like =country).sum().sum())
            var['Final Energy|Residential and Commercial|Residential|Heating|Heat pumps'] = - 3.6e-9 * 0.5 * (n.links_t.p1.filter(like ='heat pump').filter(like='services').filter(like =country).sum().sum())
            var['Final Energy|Residential and Commercial|Commercial|Heating|Electric boilers'] = - 3.6e-9 * (n.links_t.p1.filter(like ='resistive heater').filter(like='residentioal').filter(like =country).sum().sum())
            var['Final Energy|Residential and Commercial|Residential|Heating|Electric boilers'] = - 3.6e-9 * (n.links_t.p1.filter(like ='resistive heater').filter(like='services').filter(like =country).sum().sum())
            var['Final Energy|Gas|Synthetic'] = - 3.6e-9*n.links_t.p1.filter(like ='Sabatier').filter(like =country).sum().sum()
            
            """
            Capacity : Electrolysis
            """
            #MW to GW
            var['Capacity|Hydrogen|Electricity'] = 0.001*((n.links.efficiency.filter(like ='H2 Electrolysis').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='H2 Electrolysis').filter(like =country)).sum())  
            """
            Hydrogen production
            """
            #MWh to EJ
            var['Secondary Energy|Hydrogen|Electricity'] = -3.6e-9*(n.links_t.p1.filter(like ='H2 Electrolysis').filter(like =country).sum().sum())
            
            """
            Capital cost and Lifetime 
            """
            # €_2015/kW  to US$_2010/kW
            EUR2015_USD2010 = 1.11 /1.09
            ipcc2pypsa ={'Capital Cost': 'investment',
                         'Lifetime':'lifetime'} 
            for metric in ['Capital Cost', 'Lifetime']:
                factor = EUR2015_USD2010 if metric =='Capital Cost' else 1
                var[metric + ' |Electricity|Solar|PV|Rooftop PV'] = factor * costs.loc[('solar-rooftop', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Solar|PV|Utility-scale PV'] = factor * costs.loc[('solar-utility', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Solar|PV'] = 0.5* var[metric + ' |Electricity|Solar|PV|Rooftop PV'] + 0.5*var[metric + '|Electricity|Solar|PV|Utility-scale PV']
                var[metric + '|Electricity|Wind|Onshore'] = factor * costs.loc[('onwind', ipcc2pypsa[metric]), 'value']
                var[metric + '|Electricity|Wind|Offshore'] = factor * costs.loc[('offwind', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Nuclear'] = factor * costs.loc[('nuclear', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Coal|w/o CCS'] = factor * costs.loc[('coal', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Gas|w/o CCS'] = factor * costs.loc[('OCGT', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Hydro'] = factor * costs.loc[('hydro', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Storage|Pumped Hydro Storage'] = factor * costs.loc[('PHS', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Storage|Battery Capacity|Utility-scale Battery'] = factor * costs.loc[('battery storage', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Storage|Battery Capacity'] = var [metric + '|Electricity|Storage|Battery Capacity|Utility-scale Battery']
                var[metric + '|Heating|Heat pumps'] = factor * costs.loc[('decentral air-sourced heat pump', ipcc2pypsa[metric]),'value']
                var[metric + '|Heating|Electric boilers'] = factor * costs.loc[('decentral resistive heater', ipcc2pypsa[metric]),'value']
                var[metric + '|Gas|Synthetic'] = factor * costs.loc[('methanation', ipcc2pypsa[metric]),'value']
                var[metric + '|Storage|Thermal Energy Storage|Household storage'] = factor * costs.loc[('decentral water tank storage', ipcc2pypsa[metric]),'value']
                var[metric + '|Storage|Thermal Energy Storage|District heating storage'] = factor * costs.loc[('central water tank storage', ipcc2pypsa[metric]),'value']
                var[metric + '|Hydrogen|Electricity'] = factor * costs.loc[('electrolysis', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Biomass|w/o CCS'] = factor * costs.loc[('biomass EOP', ipcc2pypsa[metric]),'value']
            
            #there is a spelling error with variables overground/underground with capital/small intitial
            # the following lines can be included in the loop when the spelling mistake is corrected
            metric='Capital Cost'
            #var[metric + '|Electricity|Storage|Hydrogen Storage Capacity|overground'] = factor * costs.loc[('hydrogen storage tank', ipcc2pypsa[metric]),'value']
            #var[metric + '|Electricity|Storage|Hydrogen Storage Capacity|underground'] = factor * costs.loc[('hydrogen storage underground', ipcc2pypsa[metric]),'value']
            
            metric='Lifetime'
            #var[metric + '|Electricity|Storage|Hydrogen Storage Capacity|Overground'] = factor * costs.loc[('hydrogen storage tank', ipcc2pypsa[metric]),'value']
            #var[metric + '|Electricity|Storage|Hydrogen Storage Capacity|Underground'] = factor * costs.loc[('hydrogen storage underground', ipcc2pypsa[metric]),'value']
            
            """
            OM Cost
            """
            #US$_2010/kW·year
            factor = EUR2015_USD2010
            var['OM Cost |Electricity|Solar|PV|Rooftop PV'] = factor * 0.01*costs.loc[('solar-rooftop', 'FOM'),'value']*costs.loc[('solar-rooftop', 'investment'),'value']
            var['OM Cost|Electricity|Solar|PV|Utility-scale PV'] = factor * 0.01*costs.loc[('solar-utility', 'FOM'),'value']*costs.loc[('solar-utility', 'investment'),'value']
            var['OM Cost|Fixed|Electricity|Solar|PV'] = 0.5* var['OM Cost |Electricity|Solar|PV|Rooftop PV'] + 0.5*var['OM Cost|Electricity|Solar|PV|Utility-scale PV']
            var['OM Cost|Fixed|Electricity|Wind|Onshore'] = factor * 0.01*costs.loc[('onwind', 'FOM'),'value']*costs.loc[('onwind', 'investment'),'value']
            var['OM Cost|Fixed|Electricity|Wind|Offshore'] = factor * 0.01*costs.loc[('offwind', 'FOM'),'value']*costs.loc[('offwind', 'investment'),'value']
            var['OM Cost|Fixed|Electricity|Nuclear'] = factor * 0.01*costs.loc[('nuclear', 'FOM'),'value']*costs.loc[('nuclear', 'investment'),'value']
            var['OM Cost|Fixed|Electricity|Coal|w/o CCS'] = factor * 0.01*costs.loc[('coal', 'FOM'),'value']*costs.loc[('coal', 'investment'),'value']
            var['OM Cost|Fixed|Electricity|Gas|w/o CCS'] = factor * 0.01*costs.loc[('OCGT', 'FOM'),'value']*costs.loc[('OCGT', 'investment'),'value']
            var['OM Cost|Fixed|Electricity|Hydro'] = factor * 0.01*costs.loc[('hydro', 'FOM'),'value']*costs.loc[('hydro', 'investment'),'value']
            var['OM Cost|Electricity|Storage|Pumped Hydro Storage'] = factor * 0.01*costs.loc[('PHS', 'FOM'),'value']*costs.loc[('PHS', 'investment'),'value']
            var['OM Cost|Electricity|Storage|Battery Capacity|Utility-scale Battery'] = factor * 0.01*costs.loc[('battery storage', 'FOM'),'value']*costs.loc[('battery storage', 'investment'),'value']
            var['OM Cost|Electricity|Storage|Battery Capacity'] = var [metric + '|Electricity|Storage|Battery Capacity|Utility-scale Battery']
            #var['OM Cost|Electricity|Storage|Hydrogen Storage Capacity|Overground'] = factor * 0.01*costs.loc[('hydrogen storage tank', 'FOM'),'value']*costs.loc[('hydrogen storage tank', 'investment'),'value']
            #var['OMCost|Electricity|Storage|Hydrogen Storage Capacity|Underground'] = factor * 0.01*costs.loc[('hydrogen storage underground', 'FOM'),'value']*costs.loc[('hydrogen storage underground', 'investment'),'value']
            var['OM Cost|Heating|Heat pumps'] =factor * 0.01*costs.loc[('decentral air-sourced heat pump', 'FOM'),'value']*costs.loc[('decentral air-sourced heat pump', 'investment'),'value']
            var['OM Cost|Heating|Electric boilers'] = factor * 0.01*costs.loc[('decentral resistive heater', 'FOM'),'value']*costs.loc[('decentral resistive heater', 'investment'),'value']
            var['OM Cost|Gas|Synthetic'] = factor * 0.01*costs.loc[('methanation', 'FOM'),'value']*costs.loc[('methanation', 'investment'),'value']
            var['OM Cost|Storage|Thermal Energy Storage|Household storage'] = factor * 0.01*costs.loc[('decentral water tank storage', 'FOM'),'value']*costs.loc[('decentral water tank storage', 'investment'),'value']
            var['OM Cost|Storage|Thermal Energy Storage|District heating storage'] = factor * 0.01*costs.loc[('central water tank storage', 'FOM'),'value']*costs.loc[('central water tank storage', 'investment'),'value']
            var['OM Cost|Fixed|Hydrogen|Electricity'] = factor * 0.01*costs.loc[('electrolysis', 'FOM'),'value']*costs.loc[('electrolysis', 'investment'),'value']
            var['OM Cost|Fixed|Electricity|Biomass|w/o CCS'] = factor * 0.01*costs.loc[('biomass EOP', 'FOM'),'value']*costs.loc[('biomass EOP', 'investment'),'value']

            """
            Efficiency
            """
            var['Effciency|Heating|Heat pumps'] = costs.loc[('decentral air-sourced heat pump', 'efficiency'),'value']
            var['Efficiency|Heating|Electric boilers'] = costs.loc[('decentral resistive heater', 'efficiency'),'value']
            var['Efficiency|Gas|Synthetic'] = costs.loc[('methanation', 'efficiency'),'value']
            var['Efficiency|Hydrogen|Electricity'] = costs.loc[('electrolysis', 'efficiency'),'value']
            var['Efficiency|Electricity|Biomass|w/o CCS'] = costs.loc[('biomass EOP', 'efficiency'),'value']


            for v in var.keys():
                ro=[r for r in ds['D'] if r.value==v][0].row
                ds.cell(row=ro, column=col).value = round(var[v],3) 
                ds.cell(row=ro, column=1).value = model
                ds.cell(row=ro, column=2).value = scenario
                ds.cell(row=ro, column=3).value = iso2name[country] #region
    # add scenario name to 'meta_scenario' sheet
    ds2 = file['meta_scenario']
    ds2.cell(row=4, column=2).value = scenario
    ds2.cell(row=4, column=4).value = model
    ds2.cell(row=4, column=5).value = model_version
    ds2.cell(row=4, column=10).value = literature_reference
    ds2.cell(row=4, column=14).value = climate_target

 
        
    file.save(f"{output_folder}/IPCC_AR6_{scenario}.xlsx")

